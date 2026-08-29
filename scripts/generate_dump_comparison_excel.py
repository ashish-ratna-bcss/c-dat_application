#!/usr/bin/env python3
"""Generate Excel: dump inventory vs required Postgres tables."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "CDAT_Dump_vs_Postgres_Comparison.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
MISS_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
TITLE_FONT = Font(bold=True, size=12)


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), max_width)


DUMPS = [
    {
        "dump_file": "PDACT",
        "dump_location": "Extreme SSD / BACKUPS 10-07-2026 / PDACT",
        "dump_size": "5.2 MB",
        "mssql_db": "mssql_dump_pdact",
        "postgres_db": "PDACT_DB",
        "restore_status": "Restored",
        "tables": [
            ("PDACT_MAIN_TABLE", 1205),
        ],
    },
    {
        "dump_file": "JRMS",
        "dump_location": "Extreme SSD / BACKUPS 10-07-2026 / JRMS",
        "dump_size": "879 MB",
        "mssql_db": "mssql_dump_jrms",
        "postgres_db": "JRMS_DB",
        "restore_status": "Restored",
        "tables": [
            ("JRMS_TOTAL", 93119),
        ],
    },
    {
        "dump_file": "IRS_REPORT",
        "dump_location": "Extreme SSD / BACKUPS 10-07-2026 / IRS_REPORT",
        "dump_size": "4.4 GB",
        "mssql_db": "mssql_dump_ir",
        "postgres_db": "IR_DB",
        "restore_status": "Restored",
        "tables": [
            ("BRIEF_FACTS", 17635),
            ("DISPOSAL_OF_PROPERTY", 14427),
            ("FAMILY_HISTORY", 143866),
            ("IMAGE_TABLE", 15272),
            ("IR_PARTICULARS", 17111),
            ("LOCAL_CONTACTS_FACILITATORS", 26546),
            ("OFFENCE_DETAILS", 17309),
            ("PREVIOUS_OFFENCE_DETAILS", 29464),
            ("RELATIONSHIP_WITH_OTHER_ASSOCIATES", 37816),
            ("ROWDY SHEETERS TO CHECK", 109186),
            ("ROWDY_SHEETERS_TOTAL", 1335),
        ],
    },
    {
        "dump_file": "CDATDUPL2",
        "dump_location": "Extreme SSD / BACKUPS 10-07-2026 / CDATDUPL2",
        "dump_size": "49 GB",
        "mssql_db": "mssql_dump_cdatdupl",
        "postgres_db": "CDATDUPL_DB",
        "restore_status": "Restored",
        "tables": [
            ("CDAT_TOWERDATA_ALL", 153021777),
            ("CDAT_TOWERDATA", 45813987),
            ("CDAT_CIVILSUPPLY", 42808928),
            ("CDAT_GAS_DETAILS", 11257465),
            ("CDAT_PASSPORT", 453175),
            ("CDATSUSPECT", 91992),
            ("CDATDUPLCDATSUSPECT", 41879),
            ("CDATPHONEAREA", 33759),
            ("ALL_INDIA_CODES_LRN_CODES", 12985),
            ("COMPLETE_MO_CLASSIFICATION", 7646),
            ("MO_IMAGE_TABLE", 2514),
            ("MNC_CODES", 340),
            ("MCC_MNC", 195),
        ],
    },
]

# Required tables from team Excel (old MSSQL name -> new Postgres DB/table)
REQUIRED = [
    # CDATDUPL -> CDATDUPL_DB
    ("CDATDUPL", "CDATDUPL_DB", "ADDRESS_OTHER_STATE", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_CIVILSUPPLY", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_GAS_DETAILS", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_LICENCE", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_PASSPORT", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_PROVIDER_MASTER", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_RTA", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDAT_STATE_MASTER", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATADDRESS", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATADDRESS_OLD", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATCELLTOWERAREANEW", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATPCSUSPECT", "Not in 4 dumps", "-"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATPHONEAREA", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "CDATSUSPECT", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "COMPLETE_MO_CLASSIFICATION", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "MCC_MNC", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "MNC_CODES", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "MO_IMAGE_TABLE", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "ndps_abstract_1", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "ROWDY_SHEETER_DATA1", "CDATDUPL2", "mssql_dump_cdatdupl"),
    ("CDATDUPL", "CDATDUPL_DB", "SUSPECT_IMAGE_TABLE", "CDATDUPL2", "mssql_dump_cdatdupl"),
    # FORMS -> IR_DB
    ("FORMS", "IR_DB", "BRIEF_FACTS", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "DISPOSAL_OF_PROPERTY", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "FAMILY_HISTORY", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "IMAGE_TABLE", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "IR_PARTICULARS", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "LOCAL_CONTACTS_FACILITATORS", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "OFFENCE_DETAILS", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "PREVIOUS_OFFENCE_DETAILS", "IRS_REPORT", "mssql_dump_ir"),
    ("FORMS", "IR_DB", "RELATIONSHIP_WITH_OTHER_ASSOCIATES", "IRS_REPORT", "mssql_dump_ir"),
    # JRMS
    ("JRMS", "JRMS_DB", "JRMS_TOTAL_2012_TO_2017", "JRMS", "mssql_dump_jrms"),
    # ROWDY
    ("ROWDY_SHEETS_DATABASE", "ROWDY_SHEETS_DB", "ROWDY_SHEETER_COMPLETE_DATA", "IRS_REPORT", "mssql_dump_ir"),
    # PDACT
    ("PDACT", "PDACT_DB", "PDACT_MAIN_TABLE", "PDACT", "mssql_dump_pdact"),
]

# Build lookup: mssql_db -> {table_upper: rows}
mssql_lookup: dict[str, dict[str, int]] = {}
for d in DUMPS:
    mssql_lookup[d["mssql_db"]] = {t.upper(): rows for t, rows in d["tables"]}

ALIASES = {
    "JRMS_TOTAL_2012_TO_2017": "JRMS_TOTAL",
    "CDATCELLTOWERAREANEW": "CDAT_TOWERDATA",
    "ROWDY_SHEETER_COMPLETE_DATA": "ROWDY_SHEETERS_TOTAL",  # partial match note
}


def find_in_dump(mssql_db: str, table: str) -> tuple[str, int | None, str]:
    if mssql_db == "-":
        return "Not in any of 4 dumps", None, "Need separate CDR call-log dump (cdatpcsuspect). Not restored yet."
    tables = mssql_lookup.get(mssql_db, {})
    key = table.upper()
    if key in tables:
        return table, tables[key], "Found in dump with same table name."
    alt = ALIASES.get(table.upper()) or ALIASES.get(table)
    if alt and alt.upper() in tables:
        return alt, tables[alt.upper()], f"Exact name not found, but similar table '{alt}' exists in dump."
    # rowdy special
    if table.upper() == "ROWDY_SHEETER_COMPLETE_DATA":
        if "ROWDY SHEETERS TO CHECK" in {k.upper(): v for k, v in tables.items()}:
            for k, v in tables.items():
                if k.upper() == "ROWDY SHEETERS TO CHECK":
                    return k, v, "Exact name missing. Larger similar table in IR dump — confirm with team."
        if "ROWDY_SHEETERS_TOTAL" in tables:
            return "ROWDY_SHEETERS_TOTAL", tables["ROWDY_SHEETERS_TOTAL"], "Exact name missing. Smaller similar table in IR dump — confirm with team."
    return "-", None, "Table not found in this dump. Need to ask team for backup."


def main() -> None:
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "CDAT Migration — Dump vs Postgres Comparison"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Prepared for: checking what is available in 4 restored MSSQL dumps vs what team needs in new Postgres DBs"
    ws["A4"] = "Server: 172.16.212.229 | Date: Aug 2026"
    ws["A6"] = "4 dumps restored to MSSQL:"
    rows = [
        ("Dump file", "MSSQL database", "New Postgres DB", "Restore status"),
        ("PDACT", "mssql_dump_pdact", "PDACT_DB", "Done"),
        ("JRMS", "mssql_dump_jrms", "JRMS_DB", "Done"),
        ("IRS_REPORT", "mssql_dump_ir", "IR_DB", "Done"),
        ("CDATDUPL2", "mssql_dump_cdatdupl", "CDATDUPL_DB", "Done"),
    ]
    for i, row in enumerate(rows, start=7):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    style_header(ws, 7, 4)

    ws["A13"] = "Quick counts (required tables from team list):"
    ws["A14"] = "Total required tables:"
    ws["B14"] = len(REQUIRED)
    missing_count = 0
    partial_count = 0
    ok_count = 0
    for r in REQUIRED:
        _, _, _, dump, mdb = r
        found_name, rows_n, note = find_in_dump(mdb, r[2])
        if rows_n is None:
            missing_count += 1
        elif "similar" in note.lower() or "confirm" in note.lower():
            partial_count += 1
        else:
            ok_count += 1
    ws["A15"] = "Found in dump:"
    ws["B15"] = ok_count
    ws["A16"] = "Partial / need confirmation:"
    ws["B16"] = partial_count
    ws["A17"] = "Missing — need to ask team:"
    ws["B17"] = missing_count
    autosize(ws)

    # --- Sheet 2: What's in each dump ---
    ws2 = wb.create_sheet("Dumps and Tables")
    h2 = ["Dump file", "Dump location", "Size", "MSSQL database", "Postgres target", "Restore status", "Table in dump", "Row count", "Notes"]
    ws2.append(h2)
    style_header(ws2, 1, len(h2))
    for d in DUMPS:
        first = True
        for tbl, cnt in d["tables"]:
            note = ""
            if tbl == "CDAT_TOWERDATA":
                note = "Likely maps to CDATCELLTOWERAREANEW in Postgres"
            if tbl in ("ROWDY SHEETERS TO CHECK", "ROWDY_SHEETERS_TOTAL"):
                note = "Rowdy-related; not exact name for ROWDY_SHEETER_COMPLETE_DATA"
            ws2.append([
                d["dump_file"] if first else "",
                d["dump_location"] if first else "",
                d["dump_size"] if first else "",
                d["mssql_db"] if first else "",
                d["postgres_db"] if first else "",
                d["restore_status"] if first else "",
                tbl,
                cnt,
                note,
            ])
            first = False
    autosize(ws2)

    # --- Sheet 3: Required vs dump comparison ---
    ws3 = wb.create_sheet("Required vs Dump")
    h3 = [
        "Old MSSQL DB name (team list)",
        "New Postgres DB",
        "Required table (team list)",
        "Expected dump file",
        "MSSQL restored as",
        "Found in dump as",
        "Rows in MSSQL",
        "Status",
        "Remarks",
    ]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))
    missing_rows = []
    for old_db, pg_db, table, dump_file, mssql_db in REQUIRED:
        found_name, row_count, note = find_in_dump(mssql_db, table)
        if row_count is None:
            status = "MISSING"
            fill = MISS_FILL
            missing_rows.append((pg_db, table, dump_file, note))
        elif "similar" in note.lower() or "confirm" in note.lower():
            status = "PARTIAL"
            fill = PARTIAL_FILL
        else:
            status = "OK"
            fill = OK_FILL
        r = ws3.max_row + 1
        ws3.append([
            old_db,
            pg_db,
            table,
            dump_file,
            mssql_db,
            found_name,
            row_count if row_count is not None else "-",
            status,
            note,
        ])
        for c in range(1, len(h3) + 1):
            if c == 8:
                ws3.cell(row=r, column=c).fill = fill
    autosize(ws3)

    # --- Sheet 4: Missing tables to ask ---
    ws4 = wb.create_sheet("Missing - Ask Team")
    ws4["A1"] = "Tables required in new Postgres but NOT found in the 4 restored dumps"
    ws4["A1"].font = TITLE_FONT
    h4 = ["Postgres DB", "Required table", "Expected dump", "What to ask / remark"]
    ws4.append([])
    ws4.append(h4)
    style_header(ws4, 3, len(h4))
    for old_db, pg_db, table, dump_file, mssql_db in REQUIRED:
        found_name, row_count, note = find_in_dump(mssql_db, table)
        if row_count is None:
            ws4.append([pg_db, table, dump_file, note])
    # extra note rows
    ws4.append([])
    ws4.append(["Note", "", "", "USB also has ADDRESS_OTHER_LOCAL_03022026.BAK (734 GB) — not restored. May contain CDATADDRESS / ADDRESS_OTHER_STATE."])
    ws4.append(["Note", "", "", "CDATPCSUSPECT (call logs) is the main missing table — needs separate dump, not in these 4 files."])
    ws4.append(["Note", "", "", "ROWDY_SHEETER_COMPLETE_DATA: confirm whether to use ROWDY SHEETERS TO CHECK (109K rows) or ROWDY_SHEETERS_TOTAL (1.3K rows) from IR dump."])
    autosize(ws4)

    # --- Sheet 5: Team original layout mirror ---
    ws5 = wb.create_sheet("Team Required List")
    ws5["A1"] = "Copy of team requirement (from shared Excel)"
    ws5["A1"].font = TITLE_FONT
    blocks = [
        ("A", "CDATDUPL", "CDATDUPL_DB", [
            "ADDRESS_OTHER_STATE", "CDAT_CIVILSUPPLY", "CDAT_GAS_DETAILS", "CDAT_LICENCE",
            "CDAT_PASSPORT", "CDAT_PROVIDER_MASTER", "CDAT_RTA", "CDAT_STATE_MASTER",
            "CDATADDRESS", "CDATADDRESS_OLD", "CDATCELLTOWERAREANEW", "CDATPCSUSPECT",
            "CDATPHONEAREA", "CDATSUSPECT", "COMPLETE_MO_CLASSIFICATION", "MCC_MNC",
            "MNC_CODES", "MO_IMAGE_TABLE", "ndps_abstract_1", "ROWDY_SHEETER_DATA1", "SUSPECT_IMAGE_TABLE",
        ]),
        ("D", "FORMS", "IR_DB", [
            "BRIEF_FACTS", "DISPOSAL_OF_PROPERTY", "FAMILY_HISTORY", "IMAGE_TABLE",
            "IR_PARTICULARS", "LOCAL_CONTACTS_FACILITATORS", "OFFENCE_DETAILS",
            "PREVIOUS_OFFENCE_DETAILS", "RELATIONSHIP_WITH_OTHER_ASSOCIATES",
        ]),
    ]
    ws5["A3"] = "DATABASE NAME"
    ws5["B3"] = "TABLE NAME"
    ws5["C3"] = "Postgres DB"
    ws5["D3"] = "In dump?"
    style_header(ws5, 3, 4)
    row = 4
    for _, old, pg, tables in blocks:
        for i, t in enumerate(tables):
            req = next((x for x in REQUIRED if x[1] == pg and x[2].upper() == t.upper()), None)
            if req:
                _, row_count, _ = find_in_dump(req[4], t)
                in_dump = "Yes" if row_count else "No"
            else:
                in_dump = "?"
            ws5.cell(row=row, column=1, value=old if i == 0 else "")
            ws5.cell(row=row, column=2, value=t)
            ws5.cell(row=row, column=3, value=pg if i == 0 else "")
            ws5.cell(row=row, column=4, value=in_dump)
            row += 1
        row += 1
    # JRMS, ROWDY, PDACT on right side style
    ws5["F3"] = "DATABASE NAME"
    ws5["G3"] = "TABLE NAME"
    ws5["H3"] = "Postgres DB"
    ws5["I3"] = "In dump?"
    style_header(ws5, 3, 4)
    extra = [
        ("JRMS", "JRMS_TOTAL_2012_TO_2017", "JRMS_DB", "Yes (as JRMS_TOTAL)"),
        ("ROWDY_SHEETS_DATABASE", "ROWDY_SHEETER_COMPLETE_DATA", "ROWDY_SHEETS_DB", "Partial — confirm name"),
        ("PDACT", "PDACT_MAIN_TABLE", "PDACT_DB", "Yes"),
    ]
    for i, (old, tbl, pg, status) in enumerate(extra):
        ws5.cell(row=4 + i, column=6, value=old)
        ws5.cell(row=4 + i, column=7, value=tbl)
        ws5.cell(row=4 + i, column=8, value=pg)
        ws5.cell(row=4 + i, column=9, value=status)
    autosize(ws5)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
